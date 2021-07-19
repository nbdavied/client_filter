import openpyxl
import random
from time import strftime, localtime

wb = openpyxl.load_workbook('原始数据.xlsx')
sheet = wb['Sheet1']

rowIndex = 0
sourceData = []
for row in sheet.rows:
    if rowIndex < 1:
        rowIndex = rowIndex + 1
        continue
    rowData = []
    for cell in row:
        rowData.append(cell.value)
    print(rowData)
    sourceData.append(rowData)

wb.close()

# 读客户对照表
wbClient = openpyxl.load_workbook('客户对照.xlsx')
sheetClient = wbClient['客户']
sheetStaff = wbClient['员工']
# 存储客户对照数据
clientStaffDict = {}
rowIndex = 0
for row in sheetClient.rows:
    if rowIndex < 1:
        rowIndex = rowIndex + 1
        continue
    clientName = row[0].value
    staffName = row[1].value
    clientStaffDict[clientName] = staffName
print(clientStaffDict)
# 存储员工数据
staffs = []
for row in sheetStaff.rows:
    staffs.append(row[0].value)
print(staffs)


# 客户净流入金额字典
clientBalDict = {}
for data in sourceData:
    clientName = data[3]
    amt = data[7]
    if not clientName in clientBalDict:
        clientBalDict[clientName] = amt
    else:
        bal = clientBalDict[clientName]
        bal = bal + amt
        clientBalDict[clientName] = bal
print(clientBalDict)

# 随机指派管户经理
def randomStaff():
    rIndex = random.randint(0, len(staffs) - 1)
    return staffs[rIndex]

# 目标数据
targetData = []
for data in sourceData:
    clientName = data[3]
    if clientBalDict[clientName] >= 200000:
        target = [data[2], data[3], data[5], data[7], data[9], data[10], data[12], data[13]]
        # 关联管户经理
        if clientName in clientStaffDict:
            target.append(clientStaffDict[clientName])
        else:
            staff = randomStaff()
            target.append(staff)
            clientStaffDict[clientName] = staff
            sheetClient.append([clientName, staff])
        targetData.append(target)

wbClient.save('客户对照.xlsx')
wbClient.close()

wbTarget = openpyxl.Workbook()
ws = wbTarget.active
titles = ['客户号', '客户姓名', '交易日期', '交易金额', '交易摘要', '交易渠道',
          '对方户名', '卡种', '对接联系人', '联系结果', '录音电话留存', '是否管户']
ws.append(titles)
for target in targetData:
    ws.append(target)

currentDate = strftime('%y%m%d', localtime())
wbTarget.save(currentDate + '.xlsx')
wbTarget.close()
